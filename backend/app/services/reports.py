"""Report-center service (Phase 8 / Issue 10).

This module owns the report-task state machine and the actual Excel
generation. Public entry points:

  * :func:`create_report_task` - persist a new task in ``pending`` state
    and (optionally) schedule the background runner.

  * :func:`process_report_task` - the worker. Reads the task row, runs
    the Excel builder, and updates the task to ``completed`` or
    ``failed``. Idempotent: re-running on a ``completed`` task creates
    a new file; running on a ``generating`` task is a no-op so a
    duplicate background schedule can't double-write the row.

  * :func:`list_report_tasks` / :func:`get_report_task` /
    :func:`count_report_tasks_by_status` - paginated / detail / summary
    read helpers used by the API and the Web UI.

The filter-to-query translation lives in :func:`_build_base_query` so
the list endpoint and the Excel builder share the same selection
logic - the report is, by definition, the filtered list rendered into
a file.

The Excel writer (:func:`_write_excel`) uses ``openpyxl`` to keep the
dependency surface small. The output has three sheets:

  1. ``概览``  - one-row filter + a small count-by-risk-level table.
  2. ``汇总``  - count-by-source, count-by-day, count-by-sentiment.
  3. ``明细``  - one row per matched opinion with all the columns
                 the user can see in the opinion list.

The file path is the absolute path of the file on disk; the API
serves it via ``FileResponse`` after a permission check.
"""
from __future__ import annotations

import json
import os
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models.analysis import AnalysisResult
from app.models.datasource import DataSource, OpinionItem
from app.models.report import (
    DESCRIPTION_MAX,
    REPORT_FILE_PREFIX,
    REPORT_RISK_LEVELS,
    REPORT_STATUS_COMPLETED,
    REPORT_STATUS_FAILED,
    REPORT_STATUS_GENERATING,
    REPORT_STATUS_PENDING,
    ReportTask,
    TITLE_MAX,
)
from app.models.user import User


# Display labels used inside the Excel workbook. Kept here so the
# service is the only place that knows the Chinese column names.
RISK_LABEL = {"low": "低", "medium": "中", "high": "高", "severe": "严重"}
SENTIMENT_LABEL = {"positive": "正面", "neutral": "中性", "negative": "负面"}
RISK_LEVEL_DISPLAY_ORDER = ("severe", "high", "medium", "low")


@dataclass
class ReportInputError(Exception):
    """Raised when the request payload cannot become a report task."""

    reason: str


@dataclass
class _OpinionBatchIterable:
    """Small iterable wrapper so report generation can stream ORM rows.

    ``_write_excel`` still needs the matched count for overview metadata;
    keeping it here avoids forcing callers to materialize a list just to
    support ``len(opinions)``.
    """

    query: Any
    batch_size: int
    matched_count: int

    def __iter__(self):
        return iter(self.query.yield_per(self.batch_size))

    def __len__(self) -> int:
        return self.matched_count


# ---------- helpers ----------


def _strip(value: Optional[str]) -> str:
    return (value or "").strip()


def _normalize_filters(
    *,
    title: str = "",
    description: str = "",
    start_at: Optional[datetime] = None,
    end_at: Optional[datetime] = None,
    risk_level: str = "",
    subject_keyword: str = "",
) -> dict[str, Any]:
    """Coerce incoming filter fields into the shape we persist.

    Rejects filter combinations that the listing logic would silently
    produce empty results for (start > end, an empty-both-window
    filter that pretends to be bounded, etc.) and trims strings so
    blank values do not contaminate SQL ``LIKE`` patterns.
    """
    title = _strip(title)[:TITLE_MAX]
    description = _strip(description)[:DESCRIPTION_MAX]
    risk_level = _strip(risk_level).lower()
    if risk_level and risk_level not in REPORT_RISK_LEVELS:
        raise ReportInputError(
            f"risk_level 必须是 {sorted(REPORT_RISK_LEVELS)} 之一或留空"
        )
    subject_keyword = _strip(subject_keyword)[:128]
    if start_at and end_at and start_at > end_at:
        raise ReportInputError("start_at 不能晚于 end_at")
    return {
        "title": title,
        "description": description,
        "start_at": start_at,
        "end_at": end_at,
        "risk_level": risk_level,
        "subject_keyword": subject_keyword,
    }


def _storage_root() -> Path:
    """Return the on-disk root for report output files.

    The default is a sibling of the SQLite database (when SQLite is
    used) so a copy of the workspace is self-contained. Tests that set
    ``REPORT_STORAGE_DIR`` in the env get an isolated directory.
    """
    override = os.environ.get("REPORT_STORAGE_DIR")
    if override:
        root = Path(override)
    else:
        settings = get_settings()
        url = settings.database_url
        if url.startswith("sqlite:///"):
            db_path = Path(url[len("sqlite:///") :]).resolve()
            root = db_path.parent / REPORT_FILE_PREFIX
        else:
            root = Path.cwd() / REPORT_FILE_PREFIX
    root.mkdir(parents=True, exist_ok=True)
    return root


def _safe_filename(task_id: int, now: datetime) -> str:
    """Return a stable, collision-free filename for the task's Excel."""
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    return f"report_{task_id:08d}_{timestamp}.xlsx"


# ---------- query construction ----------


def _build_base_query(
    db: Session,
    *,
    start_at: Optional[datetime] = None,
    end_at: Optional[datetime] = None,
    risk_level: str = "",
    subject_keyword: str = "",
):
    """Return the SQLAlchemy query for opinions matching the filters.

    Used by both :func:`list_report_tasks` (to expose the matched count
    on the list response) and the Excel writer (to materialize the
    detail rows). Kept in one place so a future filter addition only
    has to be wired once.
    """
    query = db.query(OpinionItem)
    if start_at is not None:
        query = query.filter(OpinionItem.published_at >= start_at)
    if end_at is not None:
        query = query.filter(OpinionItem.published_at <= end_at)
    if risk_level:
        query = query.join(AnalysisResult, AnalysisResult.opinion_item_id == OpinionItem.id).filter(
            AnalysisResult.level == risk_level
        )
    if subject_keyword:
        like = f"%{subject_keyword}%"
        query = query.filter(
            or_(OpinionItem.title.like(like), OpinionItem.content.like(like))
        )
    return query


# ---------- creation ----------


def create_report_task(
    db: Session,
    *,
    actor: User,
    title: str = "",
    description: str = "",
    start_at: Optional[datetime] = None,
    end_at: Optional[datetime] = None,
    risk_level: str = "",
    subject_keyword: str = "",
) -> ReportTask:
    """Persist a new task in ``pending`` state.

    Returns the unflushed :class:`ReportTask` instance; the caller is
    responsible for ``db.commit()`` and for scheduling the background
    runner. The actor fields are snapshotted so the row tells a
    coherent story if the user is later disabled / deleted.
    """
    filters = _normalize_filters(
        title=title,
        description=description,
        start_at=start_at,
        end_at=end_at,
        risk_level=risk_level,
        subject_keyword=subject_keyword,
    )
    task = ReportTask(
        title=filters["title"],
        description=filters["description"],
        status=REPORT_STATUS_PENDING,
        start_at=filters["start_at"],
        end_at=filters["end_at"],
        risk_level=filters["risk_level"],
        subject_keyword=filters["subject_keyword"],
        created_by_id=actor.id,
        created_by_username=actor.username,
    )
    db.add(task)
    db.flush()
    return task


# ---------- worker ----------


def _decode_json(value: str, default):
    if not value:
        return default
    try:
        return json.loads(value)
    except (TypeError, ValueError):
        return default


def _row_payload(opinion: OpinionItem) -> dict[str, Any]:
    """Flatten an opinion + its analysis into the detail-row dict."""
    analysis = getattr(opinion, "analysis_result", None)
    source = getattr(opinion, "source", None)
    return {
        "id": opinion.id,
        "title": opinion.title or "",
        "content": opinion.content or "",
        "url": opinion.url or "",
        "author": opinion.author or "",
        "language": opinion.language or "zh",
        "published_at": opinion.published_at,
        "fetched_at": opinion.fetched_at,
        "source_id": opinion.source_id,
        "source_code": opinion.source_code,
        "source_name": source.name if source is not None else "",
        "source_type": opinion.source_type,
        "origin": opinion.origin or "",
        "analysis_status": analysis.status if analysis is not None else "pending",
        "sentiment": analysis.sentiment if analysis is not None else None,
        "confidence": analysis.confidence if analysis is not None else None,
        "provider": analysis.provider if analysis is not None else "",
        "risk_score": analysis.score if analysis is not None else None,
        "risk_level": analysis.level if analysis is not None else None,
        "error_message": analysis.error_message if analysis is not None else "",
        "factors": _decode_json(analysis.factors, default={}) if analysis is not None else {},
    }


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _autosize(ws) -> None:
    """Best-effort column auto-sizing that does not require a font engine.

    ``openpyxl`` does not ship a font engine, so we approximate by
    measuring the length of the rendered string in characters. Wide
    columns are capped to keep the workbook readable.
    """
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            text = str(cell.value)
            # CJK characters render about twice as wide; multiply so
            # the auto-fit is closer to a spreadsheet's idea of "wide".
            cjk = sum(1 for ch in text if "一" <= ch <= "鿿")
            length = cjk * 2 + (len(text) - cjk)
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max(10, max_len + 2))


def _set_detail_column_widths(ws) -> None:
    """Apply fixed widths to the detail sheet without scanning all cells."""
    widths = {
        1: 10,   # 舆情 ID
        2: 32,   # 标题
        3: 60,   # 正文
        4: 20,   # 数据源
        5: 18,   # 数据源编码
        6: 16,   # 作者
        7: 10,   # 语言
        8: 20,   # 发布时间
        9: 20,   # 抓取时间
        10: 42,  # 原文链接
        11: 12,  # 情感
        12: 12,  # 风险等级
        13: 12,  # 风险分数
        14: 12,  # 置信度
        15: 14,  # 分析状态
        16: 16,  # 分析提供方
        17: 48,  # 分析说明
        18: 36,  # 错误信息
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_excel(
    task: ReportTask,
    opinions: Iterable[OpinionItem],
    output_path: Path,
    matched_count: int,
) -> int:
    """Materialize ``opinions`` into a 3-sheet Excel workbook.

    Returns the file size in bytes so the caller can persist it on the
    task row without an extra ``stat()`` call.
    """
    wb = Workbook()

    # ---- sheet 1: 概览 (one-row filter summary) ----
    overview = wb.active
    overview.title = "概览"
    overview.append(["舆情报告 - 概览"])
    overview["A1"].font = Font(size=14, bold=True)
    overview.append([])
    overview.append(["任务 ID", task.id])
    overview.append(["任务标题", task.title or f"报告 #{task.id}"])
    overview.append(["任务说明", task.description or ""])
    overview.append(["创建人", task.created_by_username or ""])
    overview.append(["创建时间", _safe_str(task.created_at)])
    overview.append(["开始时间", _safe_str(task.start_at)])
    overview.append(["结束时间", _safe_str(task.end_at)])
    overview.append(["风险等级过滤", RISK_LABEL.get(task.risk_level, "") if task.risk_level else "全部"])
    overview.append(["关键词过滤", task.subject_keyword or "无"])
    overview.append(["匹配条数", matched_count])
    overview.append([])
    overview.append(["按风险等级统计"])
    overview[f"A{overview.max_row}"].font = _BOLD
    overview.append(["风险等级", "条数"])
    risk_header_row = overview.max_row
    overview.cell(row=risk_header_row, column=1).fill = _HEADER_FILL
    overview.cell(row=risk_header_row, column=2).fill = _HEADER_FILL
    overview.cell(row=risk_header_row, column=1).font = _HEADER_FONT
    overview.cell(row=risk_header_row, column=2).font = _HEADER_FONT
    overview.cell(row=risk_header_row, column=1).alignment = _CENTER
    overview.cell(row=risk_header_row, column=2).alignment = _CENTER

    risk_counts: Counter[str] = Counter()
    sentiment_counts: Counter[str] = Counter()
    source_counts: Counter[str] = Counter()
    day_counts: Counter[str] = Counter()
    rows: list[dict[str, Any]] = []

    for opinion in opinions:
        row = _row_payload(opinion)
        rows.append(row)
        if row["risk_level"]:
            risk_counts[row["risk_level"]] += 1
        if row["sentiment"]:
            sentiment_counts[row["sentiment"]] += 1
        if row["source_name"] or row["source_code"]:
            source_counts[row["source_name"] or row["source_code"]] += 1
        published = row["published_at"]
        if published is not None:
            day_counts[published.strftime("%Y-%m-%d")] += 1

    for level in RISK_LEVEL_DISPLAY_ORDER:
        overview.append([RISK_LABEL[level], risk_counts.get(level, 0)])

    # ---- sheet 2: 汇总 (count-by-source / day / sentiment) ----
    summary = wb.create_sheet("汇总")
    summary.append(["舆情报告 - 汇总统计"])
    summary["A1"].font = Font(size=14, bold=True)
    summary.append([])

    def _append_section(title_text: str, header: list[str], rows: Iterable[list[Any]]) -> None:
        """Append a ``title / header / rows`` block and style the header row.

        The function tracks the row of the header explicitly so empty
        sections still get a header (and so the styling does not
        silently land on a stale row from a previous block).
        """
        summary.append([title_text])
        summary.cell(row=summary.max_row, column=1).font = _BOLD
        summary.append(header)
        header_row = summary.max_row
        summary.cell(row=header_row, column=1).fill = _HEADER_FILL
        summary.cell(row=header_row, column=2).fill = _HEADER_FILL
        summary.cell(row=header_row, column=1).font = _HEADER_FONT
        summary.cell(row=header_row, column=2).font = _HEADER_FONT
        summary.cell(row=header_row, column=1).alignment = _CENTER
        summary.cell(row=header_row, column=2).alignment = _CENTER
        for row in rows:
            summary.append(row)

    _append_section(
        "按数据源统计",
        ["数据源", "条数"],
        ([name, count] for name, count in sorted(source_counts.items(), key=lambda x: (-x[1], x[0]))),
    )
    summary.append([])

    _append_section(
        "按情感倾向统计",
        ["情感", "条数"],
        ([SENTIMENT_LABEL[s], sentiment_counts.get(s, 0)] for s in ("negative", "neutral", "positive")),
    )
    summary.append([])

    _append_section(
        "按发布日期统计",
        ["日期", "条数"],
        ([day, count] for day, count in sorted(day_counts.items())),
    )

    # ---- sheet 3: 明细 (one row per opinion) ----
    detail = wb.create_sheet("明细")
    headers = [
        "舆情 ID",
        "标题",
        "正文",
        "数据源",
        "数据源编码",
        "作者",
        "语言",
        "发布时间",
        "抓取时间",
        "原文链接",
        "情感",
        "风险等级",
        "风险分数",
        "置信度",
        "分析状态",
        "分析提供方",
        "分析说明",
        "错误信息",
    ]
    detail.append(headers)
    for cell in detail[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER

    for row in rows:
        explanation_text = "\n".join(
            str(item) for item in (row["factors"].get("explanation") or [])
        ) if isinstance(row["factors"].get("explanation"), list) else ""
        detail.append(
            [
                row["id"],
                row["title"],
                row["content"],
                row["source_name"],
                row["source_code"],
                row["author"],
                row["language"],
                _safe_str(row["published_at"]),
                _safe_str(row["fetched_at"]),
                row["url"],
                SENTIMENT_LABEL.get(row["sentiment"], ""),
                RISK_LABEL.get(row["risk_level"], ""),
                row["risk_score"] if row["risk_score"] is not None else "",
                round(row["confidence"], 4) if isinstance(row["confidence"], (int, float)) else "",
                row["analysis_status"],
                row["provider"],
                explanation_text,
                row["error_message"],
            ]
        )

    # Style the detail rows: align / wrap text-heavy cells, leave
    # numeric / status cells centered.
    for r in range(2, detail.max_row + 1):
        for c in (2, 3, 10, 17, 18):  # title, content, url, explanation, error
            detail.cell(row=r, column=c).alignment = _WRAP
        for c in (1, 4, 5, 6, 7, 8, 9, 11, 12, 13, 14, 15, 16):
            detail.cell(row=r, column=c).alignment = _CENTER

    for ws in (overview, summary):
        _autosize(ws)
    _set_detail_column_widths(detail)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path.stat().st_size


def process_report_task(task_id: int) -> ReportTask:
    """Generate the Excel for ``task_id`` and update the row.

    This is the function the API schedules with ``BackgroundTasks``.
    It opens its own short-lived session because the request-scoped
    ``get_db`` dependency has been torn down by the time a
    BackgroundTask fires (FastAPI runs dependency cleanup before
    background tasks). Tests can call this directly - it does not
    require the test's session to still be open.
    """
    from app.db.session import SessionLocal

    session: Session = SessionLocal()
    try:
        task = session.get(ReportTask, task_id)
        if task is None:
            raise ReportInputError(f"report task {task_id} not found")

        # If a previous schedule already flipped the row to
        # ``generating`` (e.g. a duplicate background fire), treat the
        # second call as a no-op so we never double-write.
        if task.status == REPORT_STATUS_GENERATING:
            return task

        task.status = REPORT_STATUS_GENERATING
        task.started_at = datetime.now(timezone.utc)
        task.error_message = ""
        session.add(task)
        session.commit()

        try:
            base_query = _build_base_query(
                session,
                start_at=task.start_at,
                end_at=task.end_at,
                risk_level=task.risk_level,
                subject_keyword=task.subject_keyword,
            )
            matched_count = int(base_query.order_by(None).count())
            max_rows = max(1, int(get_settings().report_max_rows))
            if matched_count > max_rows:
                task.status = REPORT_STATUS_FAILED
                task.matched_count = matched_count
                task.included_count = 0
                task.file_path = ""
                task.file_size_bytes = 0
                task.completed_at = datetime.now(timezone.utc)
                task.error_message = (
                    f"Report matched {matched_count} rows, exceeding the maximum export "
                    f"limit of {max_rows}. Please narrow the filters and try again."
                )
                session.add(task)
                session.commit()
                return task

            ordered_query = base_query.order_by(
                OpinionItem.published_at.desc().nulls_last(),
                OpinionItem.id.desc(),
            )
            opinions = _OpinionBatchIterable(
                query=ordered_query,
                batch_size=max(1, int(get_settings().report_export_batch_size)),
                matched_count=matched_count,
            )
            # Materialize the matched count for the list endpoint up
            # front. ``included_count`` is the same number today; the
            # two columns are split so a future "cap rows" or
            # "truncate long content" change can record the
            # difference.
            task.matched_count = matched_count
            task.included_count = len(opinions)

            root = _storage_root()
            filename = _safe_filename(task.id, datetime.now(timezone.utc))
            output_path = root / filename
            size = _write_excel(task, opinions, output_path, matched_count)

            # Re-fetch after the write so the final state we
            # commit is from a fresh row, not the possibly-stale
            # ``task`` reference (the session may have been
            # implicitly expired during the long write).
            fresh = session.get(ReportTask, task_id)
            if fresh is None:
                return task
            fresh.file_path = str(output_path)
            fresh.file_size_bytes = size
            fresh.status = REPORT_STATUS_COMPLETED
            fresh.completed_at = datetime.now(timezone.utc)
            fresh.error_message = ""
            session.add(fresh)
            session.commit()
            return fresh
        except Exception as exc:  # noqa: BLE001 - we want to capture every failure
            session.rollback()
            fresh = session.get(ReportTask, task_id)
            if fresh is not None:
                fresh.status = REPORT_STATUS_FAILED
                fresh.error_message = _safe_error_message(exc)[:1000]
                fresh.completed_at = datetime.now(timezone.utc)
                fresh.file_path = ""
                fresh.file_size_bytes = 0
                session.add(fresh)
                session.commit()
            return fresh if fresh is not None else task
    finally:
        session.close()


def _safe_error_message(exc: BaseException) -> str:
    """Best-effort single-line error message that survives Excel write paths."""
    text = f"{type(exc).__name__}: {exc}"
    text = re.sub(r"\s+", " ", text).strip()
    return text[:1000]


# ---------- read helpers ----------


def _base_query_for_tasks(db: Session):
    return db.query(ReportTask)


def list_report_tasks(
    db: Session,
    *,
    status_filter: Optional[str] = None,
    creator_id: Optional[int] = None,
    limit: int = 50,
    offset: int = 0,
) -> tuple[list[ReportTask], int]:
    """Return a filtered, paginated list of report tasks."""
    query = _base_query_for_tasks(db)
    count_query = db.query(func.count(ReportTask.id))

    if status_filter:
        query = query.filter(ReportTask.status == status_filter)
        count_query = count_query.filter(ReportTask.status == status_filter)
    if creator_id is not None:
        query = query.filter(ReportTask.created_by_id == creator_id)
        count_query = count_query.filter(ReportTask.created_by_id == creator_id)

    total = int(count_query.scalar() or 0)
    rows = (
        query.order_by(ReportTask.created_at.desc(), ReportTask.id.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    return rows, total


def get_report_task(db: Session, task_id: int) -> Optional[ReportTask]:
    return db.get(ReportTask, task_id)


def count_report_tasks_by_status(db: Session) -> dict[str, int]:
    rows = (
        db.query(ReportTask.status, func.count(ReportTask.id))
        .group_by(ReportTask.status)
        .all()
    )
    counts = {
        REPORT_STATUS_PENDING: 0,
        REPORT_STATUS_GENERATING: 0,
        REPORT_STATUS_COMPLETED: 0,
        REPORT_STATUS_FAILED: 0,
    }
    for status, count in rows:
        if status in counts:
            counts[status] = int(count)
    counts["total"] = sum(
        counts[s] for s in (REPORT_STATUS_PENDING, REPORT_STATUS_GENERATING, REPORT_STATUS_COMPLETED, REPORT_STATUS_FAILED)
    )
    return counts


# ---------- serialization ----------


def report_task_to_dict(task: ReportTask) -> dict[str, Any]:
    """Flatten a row for the API / UI."""
    return {
        "id": task.id,
        "title": task.title,
        "description": task.description,
        "status": task.status,
        "error_message": task.error_message,
        "file_path": task.file_path,
        "start_at": task.start_at,
        "end_at": task.end_at,
        "risk_level": task.risk_level,
        "subject_keyword": task.subject_keyword,
        "matched_count": task.matched_count,
        "included_count": task.included_count,
        "file_size_bytes": task.file_size_bytes,
        "started_at": task.started_at,
        "completed_at": task.completed_at,
        "created_by_id": task.created_by_id,
        "created_by_username": task.created_by_username,
        "created_at": task.created_at,
        "updated_at": task.updated_at,
    }


def matched_opinions_for_task(db: Session, task: ReportTask) -> list[OpinionItem]:
    """Return the opinion rows that the report would include.

    Used by tests to verify the Excel content without needing to
    re-parse the file. Mirrors the filter logic in
    :func:`process_report_task`.
    """
    return (
        _build_base_query(
            db,
            start_at=task.start_at,
            end_at=task.end_at,
            risk_level=task.risk_level,
            subject_keyword=task.subject_keyword,
        )
        .order_by(OpinionItem.published_at.desc().nulls_last(), OpinionItem.id.desc())
        .all()
    )
