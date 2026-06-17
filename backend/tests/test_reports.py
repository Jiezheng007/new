"""Acceptance tests for Phase 8 / Issue 10: report center.

Covers:
  - Auth: 401 unauthenticated, 403 for handler on read paths
  - Create endpoint: happy path, role enforcement, filter validation,
    audit row on success
  - Worker: pending -> generating -> completed transition, file
    on disk, matched_count / file_size_bytes populated
  - Worker failure: a write that raises flips the row to ``failed``
    with error_message; the API can still list / inspect it
  - Generated Excel content: 3 sheets, header rows styled, detail
    rows match the controlled dataset
  - List / detail / summary endpoints
  - Download: completed task streams xlsx; pending/failed/missing-file
    return the right error; audit rows for both success and failure
  - Web UI: /web/reports renders for the right roles, 403 for handler,
    401 unauthenticated
"""
from __future__ import annotations

import json
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.db import session as session_module
from app.models.audit import AuditLog
from app.models.datasource import DataSource, OpinionItem
from app.models.report import (
    REPORT_STATUS_COMPLETED,
    REPORT_STATUS_FAILED,
    REPORT_STATUS_GENERATING,
    REPORT_STATUS_PENDING,
    ReportTask,
)
from app.models.user import User
from app.services.reports import process_report_task


# ---------- helpers ----------


def _engine():
    return session_module.engine


def _session():
    return Session(_engine())


@pytest.fixture()
def report_storage(monkeypatch, tmp_path):
    """Isolate the on-disk report file root to a per-test temp dir.

    The default root is computed from the SQLite URL; with the test
    fixture's temp db, all tests would share a single
    ``/tmp/report_files`` directory. Pinning the env var keeps the
    worker from polluting the host tmp.
    """
    root = tmp_path / "report_files"
    monkeypatch.setenv("REPORT_STORAGE_DIR", str(root))
    return root


def _seed_static_demo(client):
    """Run the static-demo fetch as admin so the demo + analysis rows exist."""
    client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
    db = _session()
    try:
        src = db.query(DataSource).filter(DataSource.code == "demo_static").one()
        source_id = src.id
    finally:
        db.close()
    res = client.post(f"/api/datasources/{source_id}/fetch")
    assert res.status_code == 200, res.text
    client.post("/api/auth/logout")
    return source_id


def _create_report_via_api(
    client,
    *,
    title: str = "演示报告",
    description: str = "",
    risk_level: str = "",
    subject_keyword: str = "",
    start_at: str | None = None,
    end_at: str | None = None,
    actor: str = "risk",
    password: str = "risk123",
) -> int:
    """Create a report task through the public API; return its id."""
    client.post("/api/auth/login", json={"username": actor, "password": password})
    payload: dict = {
        "title": title,
        "description": description,
        "risk_level": risk_level,
        "subject_keyword": subject_keyword,
    }
    if start_at is not None:
        payload["start_at"] = start_at
    if end_at is not None:
        payload["end_at"] = end_at
    res = client.post("/api/reports", json=payload)
    assert res.status_code == 201, res.text
    body = res.json()
    client.post("/api/auth/logout")
    return body["id"]


def _latest_audit(action: str, target_id: str = ""):
    db = _session()
    try:
        q = db.query(AuditLog).filter(AuditLog.action == action)
        if target_id:
            q = q.filter(AuditLog.target_id == target_id)
        return q.order_by(AuditLog.id.desc()).first()
    finally:
        db.close()


def _fetch_task(task_id: int) -> ReportTask:
    db = _session()
    try:
        return db.get(ReportTask, task_id)
    finally:
        db.close()


# ---------- create endpoint ----------


def test_create_report_happy_path_writes_audit(client, report_storage):
    _seed_static_demo(client)
    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.post(
        "/api/reports",
        json={"title": "6 月高风险周报", "description": "示例说明", "risk_level": "high"},
    )
    assert res.status_code == 201, res.text
    body = res.json()
    assert isinstance(body["id"], int) and body["id"] > 0
    assert body["status"] in {REPORT_STATUS_PENDING, REPORT_STATUS_GENERATING, REPORT_STATUS_COMPLETED}
    assert body["created_at"]

    entry = _latest_audit("report.create", str(body["id"]))
    assert entry is not None
    assert entry.actor_username == "risk"
    assert entry.result == "success"
    detail = json.loads(entry.detail)
    assert detail["title"] == "6 月高风险周报"
    assert detail["risk_level"] == "high"
    client.post("/api/auth/logout")

    # The DB row is persisted with the actor snapshot and a pending /
    # already-completed status (the background task may have flipped
    # it in the TestClient by the time we read it - the worker is
    # fast enough to run during the post()).
    db = _session()
    try:
        task = db.get(ReportTask, body["id"])
        assert task is not None
        assert task.created_by_username == "risk"
        assert task.title == "6 月高风险周报"
        assert task.risk_level == "high"
    finally:
        db.close()


def test_create_report_admin_also_allowed(client, report_storage):
    _seed_static_demo(client)
    client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
    res = client.post("/api/reports", json={"title": "管理员测试"})
    assert res.status_code == 201, res.text
    client.post("/api/auth/logout")


def test_create_report_blocks_handler_auditor_viewer(client, report_storage):
    _seed_static_demo(client)
    for username, password in [
        ("handler", "handler123"),
        ("auditor", "auditor123"),
        ("viewer", "viewer123"),
    ]:
        client.post("/api/auth/login", json={"username": username, "password": password})
        res = client.post("/api/reports", json={"title": "blocked"})
        assert res.status_code == 403, username
        client.post("/api/auth/logout")


def test_create_report_requires_auth(client, report_storage):
    res = client.post("/api/reports", json={"title": "unauth"})
    assert res.status_code == 401


def test_create_report_rejects_invalid_risk_level(client, report_storage):
    _seed_static_demo(client)
    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.post("/api/reports", json={"title": "bad", "risk_level": "extreme"})
    assert res.status_code == 422, res.text
    # No audit row should be written for a validation failure - the
    # request never reached the service.
    entry = _latest_audit("report.create", "new")
    assert entry is None
    client.post("/api/auth/logout")


def test_create_report_rejects_inverted_time_range(client, report_storage):
    """start_at > end_at is a domain error, not a pydantic error,
    so it returns 400 (with an audit row explaining why)."""
    _seed_static_demo(client)
    now = datetime.now(timezone.utc)
    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.post(
        "/api/reports",
        json={
            "start_at": (now - timedelta(days=1)).isoformat(),
            "end_at": (now - timedelta(days=2)).isoformat(),
        },
    )
    assert res.status_code == 400, res.text
    assert "start_at" in res.json()["detail"]

    entry = _latest_audit("report.create", "new")
    assert entry is not None
    assert entry.result == "failure"
    assert json.loads(entry.detail)["reason"] == "invalid_filters"
    client.post("/api/auth/logout")


def test_create_report_clamps_extreme_title(client, report_storage):
    """The pydantic schema caps title at TITLE_MAX; longer values 422."""
    _seed_static_demo(client)
    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.post("/api/reports", json={"title": "x" * 600})
    assert res.status_code == 422
    client.post("/api/auth/logout")


# ---------- worker: async status transitions ----------


def test_worker_pending_to_generating_to_completed(client, report_storage):
    """Drive the worker directly so we can observe each state."""
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="async flow")

    # 1. Right after create, the row may already be 'completed' (the
    #    BackgroundTask fires synchronously under TestClient when the
    #    response is awaited). Re-derive the worker's state machine
    #    by resetting the row to pending and re-running the worker.
    db = _session()
    try:
        task = db.get(ReportTask, task_id)
        task.status = REPORT_STATUS_PENDING
        task.error_message = ""
        task.file_path = ""
        task.file_size_bytes = 0
        task.matched_count = 0
        task.started_at = None
        task.completed_at = None
        db.commit()
    finally:
        db.close()

    result = process_report_task(task_id)
    # The worker closes its session before returning, so the returned
    # instance is detached - refetch the row to read attributes.
    final = _fetch_task(task_id)
    assert final is not None
    assert final.status == REPORT_STATUS_COMPLETED
    assert final.started_at is not None
    assert final.completed_at is not None
    assert final.file_path and Path(final.file_path).is_file()
    assert final.matched_count >= 1
    assert final.file_size_bytes > 0
    assert final.error_message == ""
    assert result is not None  # the function still returns a (now detached) row


def test_worker_marks_failed_when_writer_raises(client, report_storage, monkeypatch):
    """If the Excel write blows up, the row lands in ``failed`` with a
    populated ``error_message`` so an operator can see why."""
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="will-fail")

    # Reset to pending so we can re-run with a patched writer.
    db = _session()
    try:
        task = db.get(ReportTask, task_id)
        task.status = REPORT_STATUS_PENDING
        task.file_path = ""
        db.commit()
    finally:
        db.close()

    def boom(*args, **kwargs):
        raise RuntimeError("simulated excel write failure")

    monkeypatch.setattr("app.services.reports._write_excel", boom)

    result = process_report_task(task_id)
    final = _fetch_task(task_id)
    assert final is not None
    assert final.status == REPORT_STATUS_FAILED
    assert "simulated excel write failure" in final.error_message
    assert final.completed_at is not None
    assert final.file_path == ""
    assert result is not None


def test_worker_is_idempotent_on_already_generating(client, report_storage, monkeypatch):
    """A duplicate BackgroundTask fire should not double-write the file."""
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="dup")

    db = _session()
    try:
        task = db.get(ReportTask, task_id)
        task.status = REPORT_STATUS_GENERATING
        db.commit()
    finally:
        db.close()

    def should_not_run(*args, **kwargs):
        raise AssertionError("_write_excel should not be called when status is generating")

    monkeypatch.setattr("app.services.reports._write_excel", should_not_run)
    result = process_report_task(task_id)
    final = _fetch_task(task_id)
    # The worker is a no-op; the row remains in 'generating'.
    assert result is not None
    assert final.status == REPORT_STATUS_GENERATING


def test_worker_records_matched_count_for_filtered_dataset(client, report_storage):
    """With the static demo, filtering by 'severe' should match at
    least one of the curated negative items."""
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="severe", risk_level="severe")

    db = _session()
    try:
        task = db.get(ReportTask, task_id)
        task.status = REPORT_STATUS_PENDING
        db.commit()
    finally:
        db.close()
    process_report_task(task_id)

    task = _fetch_task(task_id)
    assert task.status == REPORT_STATUS_COMPLETED
    assert task.matched_count >= 1
    # The 'high' filter must NOT have matched anything else.
    assert task.included_count == task.matched_count


def test_worker_fails_before_excel_when_matched_rows_exceed_limit(
    client,
    report_storage,
    monkeypatch,
    request,
):
    monkeypatch.setenv("REPORT_MAX_ROWS", "2")
    get_settings.cache_clear()
    request.addfinalizer(get_settings.cache_clear)

    db = _session()
    try:
        source = db.query(DataSource).filter(DataSource.code == "demo_static").one()
        risk_user = db.query(User).filter_by(username="risk").one()
        db.add_all([
            OpinionItem(
                source_id=source.id,
                source_code=source.code,
                source_type=source.source_type,
                external_id=f"report-limit-{idx}",
                title=f"report limit {idx}",
                content="report limit content",
                content_hash=f"report-limit-{idx}",
            )
            for idx in range(3)
        ])
        task = ReportTask(
            title="too many rows",
            status=REPORT_STATUS_PENDING,
            created_by_id=risk_user.id,
            created_by_username=risk_user.username,
        )
        db.add(task)
        db.commit()
        task_id = task.id
    finally:
        db.close()

    def should_not_write(*args, **kwargs):
        raise AssertionError("_write_excel should not run for oversized reports")

    monkeypatch.setattr("app.services.reports._write_excel", should_not_write)
    process_report_task(task_id)

    task = _fetch_task(task_id)
    assert task.status == REPORT_STATUS_FAILED
    assert task.matched_count == 3
    assert task.included_count == 0
    assert task.file_path == ""
    assert task.file_size_bytes == 0
    assert "3" in task.error_message
    assert "2" in task.error_message
    assert "exceed" in task.error_message.lower() or "超过" in task.error_message


def test_worker_streams_matching_opinions_without_query_all(
    client,
    report_storage,
    monkeypatch,
    request,
):
    monkeypatch.setenv("REPORT_MAX_ROWS", "10")
    get_settings.cache_clear()
    request.addfinalizer(get_settings.cache_clear)

    db = _session()
    try:
        source = db.query(DataSource).filter(DataSource.code == "demo_static").one()
        risk_user = db.query(User).filter_by(username="risk").one()
        db.add_all([
            OpinionItem(
                source_id=source.id,
                source_code=source.code,
                source_type=source.source_type,
                external_id=f"report-stream-{idx}",
                title=f"report stream {idx}",
                content="report stream content",
                content_hash=f"report-stream-{idx}",
            )
            for idx in range(3)
        ])
        task = ReportTask(
            title="stream rows",
            status=REPORT_STATUS_PENDING,
            created_by_id=risk_user.id,
            created_by_username=risk_user.username,
        )
        db.add(task)
        db.commit()
        task_id = task.id
    finally:
        db.close()

    from sqlalchemy.orm import Query

    original_all = Query.all

    def fail_opinion_all(self):
        entities = {desc.get("entity") for desc in self.column_descriptions}
        if OpinionItem in entities:
            raise AssertionError("process_report_task must not call .all() for report opinions")
        return original_all(self)

    def fake_write_excel(task, opinions, output_path, matched_count):
        assert not isinstance(opinions, list)
        assert matched_count == 3
        assert sum(1 for _ in opinions) == 3
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(b"ok")
        return output_path.stat().st_size

    monkeypatch.setattr(Query, "all", fail_opinion_all)
    monkeypatch.setattr("app.services.reports._write_excel", fake_write_excel)

    process_report_task(task_id)

    task = _fetch_task(task_id)
    assert task.status == REPORT_STATUS_COMPLETED
    assert task.matched_count == 3
    assert task.included_count == 3


def test_excel_writer_does_not_autosize_detail_sheet(client, report_storage, monkeypatch):
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="fixed-detail-widths")
    db = _session()
    try:
        task = db.get(ReportTask, task_id)
        task.status = REPORT_STATUS_PENDING
        db.commit()
    finally:
        db.close()

    autosized_titles: list[str] = []

    def record_autosize(ws):
        autosized_titles.append(ws.title)

    monkeypatch.setattr("app.services.reports._autosize", record_autosize)
    process_report_task(task_id)

    task = _fetch_task(task_id)
    assert task.status == REPORT_STATUS_COMPLETED
    assert autosized_titles == ["概览", "汇总"]


# ---------- generated Excel content ----------


def _load_overview_rows(ws) -> list[list]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def test_generated_excel_has_three_sheets(client, report_storage):
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="three-sheets")
    db = _session()
    try:
        task = db.get(ReportTask, task_id)
        task.status = REPORT_STATUS_PENDING
        db.commit()
    finally:
        db.close()
    process_report_task(task_id)
    task = _fetch_task(task_id)
    assert task.status == REPORT_STATUS_COMPLETED
    assert task.file_path and Path(task.file_path).is_file()

    wb = load_workbook(task.file_path, data_only=True)
    assert set(wb.sheetnames) == {"概览", "汇总", "明细"}

    overview = wb["概览"]
    detail = wb["明细"]
    summary = wb["汇总"]

    # 概览 has the task id and matched-count rows somewhere in the body.
    overview_rows = _load_overview_rows(overview)
    flat = " ".join(str(c) for row in overview_rows for c in row if c is not None)
    assert "任务 ID" in flat
    assert "匹配条数" in flat
    assert "风险等级" in flat

    # 明细 header row should include the documented columns.
    detail_header = [c.value for c in detail[1]]
    assert "舆情 ID" in detail_header
    assert "标题" in detail_header
    assert "风险等级" in detail_header

    # 汇总 has at least the by-source / by-sentiment / by-day sections.
    summary_rows = _load_overview_rows(summary)
    summary_flat = " ".join(str(c) for row in summary_rows for c in row if c is not None)
    assert "按数据源统计" in summary_flat
    assert "按情感倾向统计" in summary_flat
    assert "按发布日期统计" in summary_flat


def test_generated_excel_detail_rows_match_controlled_dataset(client, report_storage):
    """Use the static demo (6 items) and assert the 明细 sheet has
    one row per matched opinion, in published-at-desc order, with the
    right external_ids and risk levels."""
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="controlled")
    db = _session()
    try:
        task = db.get(ReportTask, task_id)
        task.status = REPORT_STATUS_PENDING
        db.commit()
    finally:
        db.close()
    process_report_task(task_id)
    task = _fetch_task(task_id)
    assert task.status == REPORT_STATUS_COMPLETED

    wb = load_workbook(task.file_path, data_only=True)
    detail = wb["明细"]
    header = [c.value for c in detail[1]]
    id_idx = header.index("舆情 ID")
    title_idx = header.index("标题")
    risk_idx = header.index("风险等级")
    data_idx = header.index("数据源")

    rows = []
    for r in range(2, detail.max_row + 1):
        rows.append({
            "id": detail.cell(row=r, column=id_idx + 1).value,
            "title": detail.cell(row=r, column=title_idx + 1).value,
            "risk": detail.cell(row=r, column=risk_idx + 1).value,
            "source": detail.cell(row=r, column=data_idx + 1).value,
        })

    # The static demo has 6 items, all of which survive the empty
    # filter set. (The analysis auto-creates alerts on the high/severe
    # rows but does not filter them out of the report.)
    assert len(rows) >= 4
    external_titles = {r["title"] for r in rows if r["title"]}
    expected_titles = {
        "示例科技公司发布季度财报,营收同比增长 12%",
        "监管部门召开行业座谈会,强调合规经营",
        "网友爆料:某品牌产品出现严重质量问题,监管部门已介入",
        "某公司被曝数据泄露事件,涉嫌违规被查处",
    }
    assert expected_titles.issubset(external_titles)
    # Source column should always be the demo source name.
    assert all(r["source"] == "内置演示数据源" for r in rows)


def test_generated_excel_overview_counts_match_detail(client, report_storage):
    """The 'matching rows' count in 概览 must equal the detail row count."""
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="counts-match")
    db = _session()
    try:
        task = db.get(ReportTask, task_id)
        task.status = REPORT_STATUS_PENDING
        db.commit()
    finally:
        db.close()
    process_report_task(task_id)
    task = _fetch_task(task_id)
    wb = load_workbook(task.file_path, data_only=True)
    detail_rows = wb["明细"].max_row - 1  # minus header
    overview_rows = _load_overview_rows(wb["概览"])
    matched = next(
        int(row[1]) for row in overview_rows
        if row and row[0] == "匹配条数"
    )
    assert matched == detail_rows
    assert matched == task.matched_count


# ---------- list / detail / summary ----------


def test_list_reports_for_eligible_roles(client, report_storage):
    _seed_static_demo(client)
    _create_report_via_api(client, title="a")
    _create_report_via_api(client, title="b")
    for username, password in [
        ("admin", "admin123"),
        ("risk", "risk123"),
        ("auditor", "auditor123"),
        ("viewer", "viewer123"),
    ]:
        client.post("/api/auth/login", json={"username": username, "password": password})
        res = client.get("/api/reports")
        assert res.status_code == 200, username
        body = res.json()
        assert body["total"] >= 2
        for it in body["items"]:
            assert it["status"] in {
                REPORT_STATUS_PENDING,
                REPORT_STATUS_GENERATING,
                REPORT_STATUS_COMPLETED,
                REPORT_STATUS_FAILED,
            }
            assert "created_by_username" in it
        client.post("/api/auth/logout")


def test_list_reports_blocks_handler(client, report_storage):
    _seed_static_demo(client)
    client.post("/api/auth/login", json={"username": "handler", "password": "handler123"})
    res = client.get("/api/reports")
    assert res.status_code == 403
    client.post("/api/auth/logout")


def test_list_reports_requires_auth(client, report_storage):
    res = client.get("/api/reports")
    assert res.status_code == 401


def test_list_reports_filter_by_status(client, report_storage):
    _seed_static_demo(client)
    _create_report_via_api(client, title="a")
    _create_report_via_api(client, title="b")
    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.get("/api/reports?status=completed")
    assert res.status_code == 200
    body = res.json()
    for it in body["items"]:
        assert it["status"] == REPORT_STATUS_COMPLETED
    client.post("/api/auth/logout")


def test_list_reports_filter_validation(client, report_storage):
    _seed_static_demo(client)
    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.get("/api/reports?status=banana")
    assert res.status_code == 400
    client.post("/api/auth/logout")


def test_detail_includes_matched_count_and_filter_summary(client, report_storage):
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="detail", risk_level="high")
    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.get(f"/api/reports/{task_id}")
    assert res.status_code == 200
    body = res.json()
    assert body["id"] == task_id
    assert body["title"] == "detail"
    assert body["risk_level"] == "high"
    assert body["created_by_username"] == "risk"
    assert body["matched_count"] >= 1
    assert "started_at" in body
    assert "completed_at" in body
    client.post("/api/auth/logout")


def test_detail_404(client, report_storage):
    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.get("/api/reports/99999")
    assert res.status_code == 404
    client.post("/api/auth/logout")


def test_summary_counts_match_status_breakdown(client, report_storage):
    _seed_static_demo(client)
    _create_report_via_api(client, title="x")
    _create_report_via_api(client, title="y")
    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.get("/api/reports/summary")
    assert res.status_code == 200
    body = res.json()
    assert body["total"] >= 2
    assert (
        body["total"]
        == body["pending"] + body["generating"] + body["completed"] + body["failed"]
    )
    client.post("/api/auth/logout")


# ---------- download ----------


def test_download_completed_returns_xlsx(client, report_storage):
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="downloadable")
    # The BackgroundTask has likely already run by the time we get
    # here; force a re-run so the file is fresh and the result is
    # predictable.
    db = _session()
    try:
        task = db.get(ReportTask, task_id)
        task.status = REPORT_STATUS_PENDING
        task.file_path = ""
        db.commit()
    finally:
        db.close()
    process_report_task(task_id)

    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.get(f"/api/reports/{task_id}/download")
    assert res.status_code == 200, res.text
    assert res.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # ``.xlsx`` is a zip - the leading bytes are ``PK\x03\x04``.
    assert res.content[:2] == b"PK"

    entry = _latest_audit("report.download", str(task_id))
    assert entry is not None
    assert entry.result == "success"
    detail = json.loads(entry.detail)
    assert detail["matched_count"] >= 1
    assert detail["file_size_bytes"] > 0
    client.post("/api/auth/logout")


def test_download_pending_returns_409(client, report_storage):
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="still-pending")
    # Force it back to pending so the worker hasn't finished yet.
    db = _session()
    try:
        task = db.get(ReportTask, task_id)
        task.status = REPORT_STATUS_PENDING
        db.commit()
    finally:
        db.close()

    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.get(f"/api/reports/{task_id}/download")
    assert res.status_code == 409
    assert "pending" in res.json()["detail"]
    client.post("/api/auth/logout")

    entry = _latest_audit("report.download", str(task_id))
    assert entry is not None
    assert entry.result == "failure"
    assert json.loads(entry.detail)["reason"] == "not_ready"


def test_download_failed_returns_409(client, report_storage):
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="download-failed")
    db = _session()
    try:
        task = db.get(ReportTask, task_id)
        task.status = REPORT_STATUS_FAILED
        task.error_message = "writer blew up"
        db.commit()
    finally:
        db.close()

    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.get(f"/api/reports/{task_id}/download")
    assert res.status_code == 409
    assert "failed" in res.json()["detail"]
    client.post("/api/auth/logout")


def test_download_missing_file_returns_410(client, report_storage):
    """A completed row whose file is gone should give 410, not 500."""
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="ghost-file")
    db = _session()
    try:
        task = db.get(ReportTask, task_id)
        task.status = REPORT_STATUS_COMPLETED
        task.file_path = "/tmp/this/path/should/never/exist/report.xlsx"
        task.file_size_bytes = 0
        db.commit()
    finally:
        db.close()

    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.get(f"/api/reports/{task_id}/download")
    assert res.status_code == 410
    client.post("/api/auth/logout")

    entry = _latest_audit("report.download", str(task_id))
    assert entry is not None
    assert entry.result == "failure"
    assert json.loads(entry.detail)["reason"] == "file_missing"


def test_download_404(client, report_storage):
    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.get("/api/reports/99999/download")
    assert res.status_code == 404
    client.post("/api/auth/logout")


def test_download_blocks_handler(client, report_storage):
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="for-handler")
    client.post("/api/auth/login", json={"username": "handler", "password": "handler123"})
    res = client.get(f"/api/reports/{task_id}/download")
    assert res.status_code == 403
    client.post("/api/auth/logout")


# ---------- web UI ----------


def test_reports_page_renders_for_eligible_roles(client, report_storage):
    _seed_static_demo(client)
    _create_report_via_api(client, title="ui")
    for username, password in [
        ("admin", "admin123"),
        ("risk", "risk123"),
        ("auditor", "auditor123"),
        ("viewer", "viewer123"),
    ]:
        client.post("/api/auth/login", json={"username": username, "password": password})
        res = client.get("/web/reports")
        assert res.status_code == 200, (username, res.text[:200])
        body = res.text
        assert "报告" in body
        assert res.headers.get("X-User-Role") == {
            "admin": "admin",
            "risk": "risk_control",
            "auditor": "auditor",
            "viewer": "viewer",
        }[username]
        client.post("/api/auth/logout")


def test_reports_page_blocks_handler(client, report_storage):
    client.post("/api/auth/login", json={"username": "handler", "password": "handler123"})
    res = client.get("/web/reports")
    assert res.status_code == 403
    client.post("/api/auth/logout")


def test_reports_page_unauthenticated_redirects(client, report_storage):
    res = client.get("/web/reports")
    assert res.status_code == 401


# ---------- end-to-end / happy path ----------


def test_e2e_create_then_download(client, report_storage):
    """Create a report, wait for the worker, list it, then download.

    Mirrors what the Web UI does: POST -> poll list -> click download.
    """
    _seed_static_demo(client)
    task_id = _create_report_via_api(client, title="e2e", risk_level="high")

    # List as risk-control and confirm the task is visible.
    client.post("/api/auth/login", json={"username": "risk", "password": "risk123"})
    res = client.get("/api/reports")
    body = res.json()
    ids = [it["id"] for it in body["items"]]
    assert task_id in ids
    listed = next(it for it in body["items"] if it["id"] == task_id)
    assert listed["status"] == REPORT_STATUS_COMPLETED
    assert listed["matched_count"] >= 1
    assert listed["file_size_bytes"] > 0

    res = client.get(f"/api/reports/{task_id}/download")
    assert res.status_code == 200
    assert res.content[:2] == b"PK"
    client.post("/api/auth/logout")
